import openpyxl

#it is used for writing same data in all rows and columns
# file = "E:\writing.xlsx"
# workbook = openpyxl.load_workbook(file)
# print(f"You rile is ready to use: {workbook}")
# sheet = workbook['Sheet1']
# # sheet = workbook.active #get active sheet from excel
# sheet = workbook.active
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "Bishu"
# workbook.save(file)

#creating multiple rows and columns with different names

file = "E:\writing.xlsx"
workbook = openpyxl.load_workbook(file)
print(f"You rile is ready to use: {workbook}")
sheet = workbook['Sheet1']
# sheet = workbook.active #get active sheet from excel
sheet = workbook.active

sheet.cell(1,1).value = "Bishu"
sheet.cell(1,2).value = "pinnika"
sheet.cell(2,1).value = '3'
sheet.cell(2,2).value = '4'
workbook.save(file)
