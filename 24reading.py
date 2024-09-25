import openpyxl


file = "E:\data.xlsx"
workbook = openpyxl.load_workbook(file)
print(f"You rile is ready to use: {workbook}")
sheet = workbook['Sheet2']

rows = sheet.max_row #read count no of rows
cols = sheet.max_column #count no of columns\

#Reading all the rows and columns from the excel sheet,
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end='          ')
    print()